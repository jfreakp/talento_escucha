from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import UserCreationForm, SetPasswordForm
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator
from tickets.models import Ticket, Agencia, TicketAuditoria
from .decorators import require_role, user_can_manage_users, user_has_any_role
from .forms import UserProfileForm, CustomPasswordChangeForm
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus.flowables import PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@login_required
def dashboard_view(request):
    """Vista principal del dashboard"""
    # Obtener estadísticas básicas
    total_tickets = Ticket.objects.count()
    total_usuarios = User.objects.count()
    total_agencias = Agencia.objects.count()
    
    # Estadísticas por estado
    tickets_pendientes = Ticket.objects.filter(estado='pendiente').count()
    tickets_en_proceso = Ticket.objects.filter(estado='en_proceso').count()
    tickets_resueltos = Ticket.objects.filter(estado='resuelto').count()
    tickets_cerrados = Ticket.objects.filter(estado='cerrado').count()
    tickets_cancelados = Ticket.objects.filter(estado='cancelado').count()
    
    # Estadísticas por tipo de petición
    peticiones = Ticket.objects.filter(tipo_solicitud='P').count()
    quejas = Ticket.objects.filter(tipo_solicitud='Q').count()
    reclamos = Ticket.objects.filter(tipo_solicitud='R').count()
    solicitudes = Ticket.objects.filter(tipo_solicitud='S').count()
    
    # Estadísticas por severidad
    severidad_alta = Ticket.objects.filter(severidad='A').count()
    severidad_media = Ticket.objects.filter(severidad='M').count()
    severidad_baja = Ticket.objects.filter(severidad='B').count()
    
    # Estadísticas por agencia (top 6)
    from django.db.models import Count
    tickets_por_agencia = Ticket.objects.values('agencia__nombre').annotate(
        count=Count('id')
    ).order_by('-count')[:6]
    
    # Tickets nuevos del día
    from datetime import date
    tickets_hoy = Ticket.objects.filter(fecha_creacion__date=date.today()).count()
    
    # Estadísticas de anonimato
    # Tickets de usuarios registrados (tienen usuario_crea)
    tickets_registrados = Ticket.objects.filter(usuario_crea__isnull=False).count()
    # Tickets anónimos (no tienen usuario_crea)
    tickets_anonimos = Ticket.objects.filter(usuario_crea__isnull=True).count()
    
    # Últimos tickets creados
    ultimos_tickets = Ticket.objects.order_by('-fecha_creacion')[:5]
    
    context = {
        'total_tickets': total_tickets,
        'total_usuarios': total_usuarios,
        'total_agencias': total_agencias,
        # Estadísticas por estado
        'tickets_pendientes': tickets_pendientes,
        'tickets_en_proceso': tickets_en_proceso,
        'tickets_resueltos': tickets_resueltos,
        'tickets_cerrados': tickets_cerrados,
        'tickets_cancelados': tickets_cancelados,
        # Estadísticas por tipo
        'peticiones': peticiones,
        'quejas': quejas,
        'reclamos': reclamos,
        'solicitudes': solicitudes,
        # Estadísticas por severidad
        'severidad_alta': severidad_alta,
        'severidad_media': severidad_media,
        'severidad_baja': severidad_baja,
        # Estadísticas adicionales
        'tickets_por_agencia': tickets_por_agencia,
        'agencias_top': tickets_por_agencia[:6],  # Top 6 agencias
        'agencias_max': tickets_por_agencia[0]['count'] if tickets_por_agencia else 1,  # Para calcular porcentajes
        'tickets_hoy': tickets_hoy,
        # Estadísticas de anonimato
        'tickets_registrados': tickets_registrados,
        'tickets_anonimos': tickets_anonimos,
        'ultimos_tickets': ultimos_tickets,
        'user': request.user,
    }
    
    return render(request, 'admin_dashboard/dashboard.html', context)


@login_required
def administracion_usuarios(request):
    """Vista para administrar usuarios"""
    # Verificar permisos
    if not user_can_manage_users(request.user):
        raise PermissionDenied("No tienes permisos para gestionar usuarios")
    
    # Obtener todos los usuarios
    usuarios = User.objects.all().order_by('-date_joined')
    
    context = {
        'usuarios': usuarios,
        'total_usuarios': usuarios.count(),
    }
    
    return render(request, 'admin_dashboard/administracion_usuarios.html', context)


@login_required
def agregar_usuario(request):
    """Vista para agregar nuevos usuarios"""
    # Verificar permisos
    if not user_can_manage_users(request.user):
        raise PermissionDenied("No tienes permisos para agregar usuarios")
    
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Agregar información adicional si se proporciona
            if request.POST.get('first_name'):
                user.first_name = request.POST.get('first_name')
            if request.POST.get('last_name'):
                user.last_name = request.POST.get('last_name')
            if request.POST.get('email'):
                user.email = request.POST.get('email')
            user.save()
            
            # Asignar rol seleccionado
            role_name = request.POST.get('role')
            if role_name:
                try:
                    role_group = Group.objects.get(name=role_name)
                    user.groups.add(role_group)
                    messages.success(request, f'Usuario {user.username} creado exitosamente con rol {role_name}.')
                except Group.DoesNotExist:
                    messages.warning(request, f'Usuario {user.username} creado, pero el rol {role_name} no existe.')
            else:
                messages.success(request, f'Usuario {user.username} creado exitosamente.')
            
            return redirect('admin_dashboard:administracion_usuarios')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = UserCreationForm()
    
    # Obtener todos los roles disponibles
    roles = Group.objects.all()
    
    context = {
        'form': form,
        'roles': roles,
    }
    
    return render(request, 'admin_dashboard/agregar_usuario.html', context)


@login_required
def eliminar_usuario(request, user_id):
    """Vista para eliminar usuarios"""
    # Verificar permisos
    if not user_can_manage_users(request.user):
        raise PermissionDenied("No tienes permisos para eliminar usuarios")
    
    usuario = get_object_or_404(User, id=user_id)
    
    # Verificar que no se está tratando de eliminar a sí mismo
    if usuario.id == request.user.id:
        messages.error(request, 'No puedes eliminar tu propia cuenta.')
        return redirect('admin_dashboard:administracion_usuarios')
    
    # Verificar que no es el superusuario
    if usuario.is_superuser:
        messages.error(request, 'No se puede eliminar una cuenta de superusuario.')
        return redirect('admin_dashboard:administracion_usuarios')
    
    if request.method == 'POST':
        username = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{username}" eliminado exitosamente.')
        return redirect('admin_dashboard:administracion_usuarios')
    
    # Si es GET, mostrar confirmación
    context = {
        'usuario': usuario,
    }
    
    return render(request, 'admin_dashboard/confirmar_eliminacion.html', context)


@login_required
def editar_usuario(request, user_id):
    """Vista para editar usuarios existentes"""
    # Verificar permisos
    if not user_can_manage_users(request.user):
        raise PermissionDenied("No tienes permisos para editar usuarios")
    
    usuario = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        # Actualizar información del usuario
        usuario.first_name = request.POST.get('first_name', '')
        usuario.last_name = request.POST.get('last_name', '')
        usuario.email = request.POST.get('email', '')
        
        # Cambiar estado activo/inactivo
        usuario.is_active = 'is_active' in request.POST
        
        # Actualizar rol
        role_name = request.POST.get('role')
        if role_name:
            # Remover todos los roles actuales
            usuario.groups.clear()
            # Asignar nuevo rol
            try:
                role_group = Group.objects.get(name=role_name)
                usuario.groups.add(role_group)
            except Group.DoesNotExist:
                messages.warning(request, f'El rol {role_name} no existe.')
        
        usuario.save()
        messages.success(request, f'Usuario {usuario.username} actualizado exitosamente.')
        return redirect('admin_dashboard:administracion_usuarios')
    
    # Obtener todos los roles disponibles
    roles = Group.objects.all()
    # Obtener rol actual del usuario
    current_role = usuario.groups.first()
    
    context = {
        'usuario': usuario,
        'roles': roles,
        'current_role': current_role,
    }
    
    return render(request, 'admin_dashboard/editar_usuario.html', context)


@login_required
@require_role('ADMIN')
def resetear_password(request):
    """Vista para mostrar la lista de usuarios para resetear password"""
    usuarios = User.objects.all().exclude(id=request.user.id)  # Excluir al usuario actual
    
    context = {
        'usuarios': usuarios,
    }
    
    return render(request, 'admin_dashboard/resetear_password.html', context)


@login_required
@require_role('ADMIN')
def resetear_password_usuario(request, user_id):
    """Vista para resetear la password de un usuario específico"""
    usuario = get_object_or_404(User, id=user_id)
    
    # No permitir resetear la password del propio usuario
    if usuario.id == request.user.id:
        messages.error(request, 'No puedes resetear tu propia contraseña.')
        return redirect('admin_dashboard:resetear_password')
    
    if request.method == 'POST':
        form = SetPasswordForm(usuario, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Contraseña reseteada exitosamente para el usuario {usuario.username}.')
            return redirect('admin_dashboard:resetear_password')
        else:
            # Si hay errores en el formulario, se mostrarán en el template
            pass
    else:
        form = SetPasswordForm(usuario)
    
    context = {
        'usuario': usuario,
        'form': form,
    }
    
    return render(request, 'admin_dashboard/resetear_password_usuario.html', context)


@login_required
def buscar_ticket(request):
    """Vista para buscar tickets por código"""
    ticket = None
    codigo_buscado = None
    error_message = None
    
    if request.method == 'POST':
        codigo_buscado = request.POST.get('codigo', '').strip().upper()
        
        if codigo_buscado:
            try:
                # Buscar el ticket por código
                ticket = Ticket.objects.get(codigo=codigo_buscado)
                # Redirigir a la vista del ticket
                return redirect('admin_dashboard:view_ticket', codigo=codigo_buscado)
            except Ticket.DoesNotExist:
                error_message = f'No se encontró ningún ticket con el código "{codigo_buscado}"'
        else:
            error_message = 'Por favor ingresa un código de ticket'
    
    context = {
        'ticket': ticket,
        'codigo_buscado': codigo_buscado,
        'error_message': error_message,
    }
    
    return render(request, 'admin_dashboard/buscar_ticket.html', context)


@login_required
def view_ticket(request, codigo):
    """Vista para mostrar el detalle del ticket con tabs"""
    ticket = get_object_or_404(Ticket, codigo=codigo)
    
    # Obtener el historial de auditoría del ticket
    historial = ticket.auditorias.all().order_by('-fecha_cambio')
    
    context = {
        'ticket': ticket,
        'historial': historial,
    }
    
    return render(request, 'admin_dashboard/view_ticket.html', context)


@login_required
def tickets_pendientes(request):
    """Vista para mostrar tickets pendientes - solo para revisores"""
    # Verificar que el usuario tenga el rol de REVISOR
    if not request.user.groups.filter(name='REVISOR').exists():
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('admin_dashboard:dashboard')
    
    # Obtener tickets sin asignar (sin usuario_asignado)
    tickets_list = Ticket.objects.filter(
        usuario_asignado__isnull=True
    ).order_by('-fecha_creacion')
    
    # Calcular estadísticas por estado de tickets sin asignar
    stats = {
        'total': tickets_list.count(),
        'pendiente': tickets_list.filter(estado='pendiente').count(),
        'en_proceso': tickets_list.filter(estado='en_proceso').count(),
        'resuelto': tickets_list.filter(estado='resuelto').count(),
        'cerrado': tickets_list.filter(estado='cerrado').count(),
        'cancelado': tickets_list.filter(estado='cancelado').count(),
    }
    
    # Implementar paginación
    paginator = Paginator(tickets_list, 10)  # 10 tickets por página
    page_number = request.GET.get('page')
    tickets = paginator.get_page(page_number)
    
    context = {
        'tickets': tickets,
        'page_title': 'Tickets Pendientes',
        'page_description': 'Lista de tickets sin asignar',
        'total_tickets': stats['total'],
        'stats': stats,
    }
    
    return render(request, 'admin_dashboard/tickets_pendientes.html', context)


@login_required
def tickets_asignados(request):
    """Vista para mostrar tickets asignados al revisor actual"""
    # Verificar que el usuario tenga el rol de REVISOR
    if not request.user.groups.filter(name='REVISOR').exists():
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('admin_dashboard:dashboard')
    
    # Obtener tickets asignados al usuario actual
    tickets_list = Ticket.objects.filter(
        usuario_asignado=request.user
    ).order_by('-fecha_creacion')
    
    # Calcular estadísticas por estado
    stats = {
        'total': tickets_list.count(),
        'pendiente': tickets_list.filter(estado='pendiente').count(),
        'en_proceso': tickets_list.filter(estado='en_proceso').count(),
        'resuelto': tickets_list.filter(estado='resuelto').count(),
        'cerrado': tickets_list.filter(estado='cerrado').count(),
        'cancelado': tickets_list.filter(estado='cancelado').count(),
    }
    
    # Implementar paginación
    paginator = Paginator(tickets_list, 10)  # 10 tickets por página
    page_number = request.GET.get('page')
    tickets = paginator.get_page(page_number)
    
    context = {
        'tickets': tickets,
        'page_title': 'Mis Tickets',
        'page_description': 'Lista de tickets asignados a mí',
        'total_tickets': stats['total'],
        'stats': stats,
    }
    
    return render(request, 'admin_dashboard/tickets_asignados.html', context)


@login_required
def asignar_ticket_a_mi(request, ticket_id):
    """Vista para que un revisor se asigne un ticket a sí mismo"""
    # Verificar que el usuario tenga el rol de REVISOR
    if not request.user.groups.filter(name='REVISOR').exists():
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('admin_dashboard:dashboard')
    
    # Verificar que sea una petición POST
    if request.method != 'POST':
        messages.error(request, "Método no permitido.")
        return redirect('admin_dashboard:tickets_pendientes')
    
    try:
        # Obtener el ticket
        ticket = get_object_or_404(Ticket, id=ticket_id)
        
        # Verificar que el ticket no esté ya asignado
        if ticket.usuario_asignado:
            messages.warning(request, f"El ticket #{ticket.codigo} ya está asignado a {ticket.usuario_asignado.get_full_name() or ticket.usuario_asignado.username}.")
            return redirect('admin_dashboard:tickets_pendientes')
        
        # Asignar el ticket al usuario actual
        ticket.usuario_asignado = request.user
        ticket.estado = 'en_proceso'  # Cambiar estado a en proceso
        ticket.usuario_actualiza = request.user
        ticket.save()  # La auditoría se registra automáticamente con el signal
        
        messages.success(request, f"Te has asignado exitosamente el ticket #{ticket.codigo}.")
        return redirect('admin_dashboard:tickets_asignados')
        
    except Exception as e:
        messages.error(request, f"Error al asignar el ticket: {str(e)}")
        return redirect('admin_dashboard:tickets_pendientes')


@login_required 
def formulario_solucion(request, ticket_id):
    """Vista para mostrar el formulario de solución de un ticket"""
    # Verificar que el usuario tenga el rol de REVISOR
    if not request.user.groups.filter(name='REVISOR').exists():
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('admin_dashboard:dashboard')
    
    # Obtener el ticket
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Verificar que el ticket esté asignado al usuario actual
    if ticket.usuario_asignado != request.user:
        messages.warning(request, f"El ticket #{ticket.codigo} no está asignado a tu usuario.")
        return redirect('admin_dashboard:tickets_asignados')
    
    # Verificar que el ticket no esté ya resuelto o cerrado
    if ticket.estado in ['resuelto', 'cerrado']:
        messages.warning(request, f"El ticket #{ticket.codigo} ya está {ticket.get_estado_display().lower()}.")
        return redirect('admin_dashboard:tickets_asignados')
    
    context = {
        'ticket': ticket,
        'page_title': f'Resolver Ticket #{ticket.codigo}',
        'page_description': 'Proporciona la solución implementada para este ticket'
    }
    
    return render(request, 'admin_dashboard/resolver_ticket.html', context)


@login_required
def resolver_ticket(request, ticket_id):
    """Vista para que un revisor agregue la solución a un ticket asignado"""
    # Verificar que el usuario tenga el rol de REVISOR
    if not request.user.groups.filter(name='REVISOR').exists():
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('admin_dashboard:dashboard')
    
    # Verificar que sea una petición POST
    if request.method != 'POST':
        messages.error(request, "Método no permitido.")
        return redirect('admin_dashboard:tickets_asignados')
    
    try:
        # Obtener el ticket
        ticket = get_object_or_404(Ticket, id=ticket_id)
        
        # Verificar que el ticket esté asignado al usuario actual
        if ticket.usuario_asignado != request.user:
            messages.warning(request, f"El ticket #{ticket.codigo} no está asignado a tu usuario.")
            return redirect('admin_dashboard:tickets_asignados')
        
        # Verificar que el ticket no esté ya resuelto o cerrado
        if ticket.estado in ['resuelto', 'cerrado']:
            messages.warning(request, f"El ticket #{ticket.codigo} ya está {ticket.get_estado_display().lower()}.")
            return redirect('admin_dashboard:tickets_asignados')
        
        # Obtener la solución del formulario
        solucion = request.POST.get('solucion', '').strip()
        if not solucion:
            messages.error(request, "Debes proporcionar una solución para el ticket.")
            return redirect('admin_dashboard:tickets_asignados')
        
        # Actualizar el ticket con la solución
        ticket.solucion = solucion
        ticket.estado = 'resuelto'
        ticket.usuario_actualiza = request.user
        ticket.save()  # La auditoría se registra automáticamente con el signal
        
        messages.success(request, f"Has marcado como resuelto el ticket #{ticket.codigo} exitosamente.")
        return redirect('admin_dashboard:tickets_asignados')
        
    except Exception as e:
        messages.error(request, f"Error al resolver el ticket: {str(e)}")
        return redirect('admin_dashboard:tickets_asignados')


@login_required
@user_has_any_role(['ADMIN', 'REVISOR'])
def reporte_tickets(request):
    """Vista para generar reportes de tickets en PDF"""
    
    # Función auxiliar para aplicar filtros
    def aplicar_filtros(query, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia):
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                query = query.filter(fecha_creacion__date__gte=fecha_desde_obj)
            except ValueError:
                pass
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                query = query.filter(fecha_creacion__date__lte=fecha_hasta_obj)
            except ValueError:
                pass
        
        if estado and estado != 'todos':
            query = query.filter(estado=estado)
        
        if tipo_solicitud and tipo_solicitud != 'todos':
            query = query.filter(tipo_solicitud=tipo_solicitud)
        
        if agencia and agencia != 'todas':
            query = query.filter(agencia_id=agencia)
        
        return query.order_by('-fecha_creacion')
    
    # Obtener filtros de POST o GET (para paginación)
    if request.method == 'POST' or request.GET.get('accion') == 'buscar':
        # Obtener filtros del formulario (POST) o parámetros URL (GET)
        if request.method == 'POST':
            fecha_desde = request.POST.get('fecha_desde')
            fecha_hasta = request.POST.get('fecha_hasta')
            estado = request.POST.get('estado')
            tipo_solicitud = request.POST.get('tipo_solicitud')
            agencia = request.POST.get('agencia')
            accion = request.POST.get('accion', 'buscar')
        else:  # GET con parámetros de búsqueda
            fecha_desde = request.GET.get('fecha_desde')
            fecha_hasta = request.GET.get('fecha_hasta')
            estado = request.GET.get('estado')
            tipo_solicitud = request.GET.get('tipo_solicitud')
            agencia = request.GET.get('agencia')
            accion = request.GET.get('accion', 'buscar')
        
        # Construir queryset base
        tickets_query = Ticket.objects.all()
        
        # Aplicar filtros
        tickets = aplicar_filtros(tickets_query, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia)
        
        # Manejar diferentes tipos de descarga
        if accion == 'descargar':
            return generar_pdf_reporte(tickets, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia)
        elif accion == 'descargar_excel':
            return generar_excel_reporte(tickets, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia)
        elif accion == 'descargar_txt':
            return generar_txt_reporte(tickets, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia)
        
        # Si la acción es buscar, mostrar resultados en pantalla
        else:
            # Paginación
            paginator = Paginator(tickets, 10)  # 10 tickets por página
            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)
            
            # Preparar información de filtros
            filtros_aplicados = {
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'estado': estado,
                'tipo_solicitud': tipo_solicitud,
                'agencia': agencia,
            }
            
            context = {
                'user': request.user,
                'estados_choices': Ticket.ESTADO_CHOICES,
                'tipos_choices': Ticket.TIPO_SOLICITUD_CHOICES,
                'agencias': Agencia.objects.all().order_by('nombre'),
                'tickets': page_obj,
                'total_tickets': tickets.count(),
                'filtros_aplicados': filtros_aplicados,
                'mostrar_resultados': True,
            }
            
            return render(request, 'admin_dashboard/reporte_tickets.html', context)
    
    # GET request - mostrar formulario
    context = {
        'user': request.user,
        'estados_choices': Ticket.ESTADO_CHOICES,
        'tipos_choices': Ticket.TIPO_SOLICITUD_CHOICES,
        'agencias': Agencia.objects.all().order_by('nombre'),
        'mostrar_resultados': False,
    }
    
    return render(request, 'admin_dashboard/reporte_tickets.html', context)


def generar_pdf_reporte(tickets, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia=None):
    """Función auxiliar para generar el PDF del reporte"""
    
    # Generar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_tickets_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=1*inch)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_LEFT
    )
    
    # Contenido del PDF
    content = []
    
    # Título
    title = Paragraph("REPORTE DE TICKETS", title_style)
    content.append(title)
    content.append(Spacer(1, 20))
    
    # Información de filtros aplicados
    filtros_info = []
    if fecha_desde:
        filtros_info.append(f"Fecha desde: {fecha_desde}")
    if fecha_hasta:
        filtros_info.append(f"Fecha hasta: {fecha_hasta}")
    if estado and estado != 'todos':
        estado_display = dict(Ticket.ESTADO_CHOICES).get(estado, estado)
        filtros_info.append(f"Estado: {estado_display}")
    if tipo_solicitud and tipo_solicitud != 'todos':
        tipo_display = dict(Ticket.TIPO_SOLICITUD_CHOICES).get(tipo_solicitud, tipo_solicitud)
        filtros_info.append(f"Tipo: {tipo_display}")
    if agencia and agencia != 'todas':
        try:
            agencia_obj = Agencia.objects.get(id=agencia)
            filtros_info.append(f"Agencia: {agencia_obj.nombre}")
        except Agencia.DoesNotExist:
            pass
    
    if filtros_info:
        filtros_text = " | ".join(filtros_info)
        filtros_para = Paragraph(f"<b>Filtros aplicados:</b> {filtros_text}", styles['Normal'])
        content.append(filtros_para)
        content.append(Spacer(1, 20))
    
    # Resumen estadístico
    total_tickets = tickets.count()
    resumen_para = Paragraph(f"<b>Total de tickets encontrados:</b> {total_tickets}", subtitle_style)
    content.append(resumen_para)
    content.append(Spacer(1, 10))
    
    # Fecha de generación
    fecha_generacion = timezone.now().strftime("%d/%m/%Y %H:%M:%S")
    fecha_para = Paragraph(f"<b>Fecha de generación:</b> {fecha_generacion}", styles['Normal'])
    content.append(fecha_para)
    content.append(Spacer(1, 20))
    
    if tickets.exists():
        # Crear tabla con los datos
        data = [['Código', 'Tipo', 'Estado', 'Severidad', 'Fecha Creación', 'Asignado a']]
        
        for ticket in tickets:
            # Formatear datos
            tipo_display = dict(Ticket.TIPO_SOLICITUD_CHOICES).get(ticket.tipo_solicitud, ticket.tipo_solicitud)
            estado_display = dict(Ticket.ESTADO_CHOICES).get(ticket.estado, ticket.estado)
            severidad_display = dict(Ticket.SEVERIDAD_CHOICES).get(ticket.severidad, ticket.severidad)
            fecha_creacion = ticket.fecha_creacion.strftime("%d/%m/%Y")
            asignado = ticket.usuario_asignado.get_full_name() if ticket.usuario_asignado else 'Sin asignar'
            
            data.append([
                ticket.codigo,
                tipo_display,
                estado_display,
                severidad_display,
                fecha_creacion,
                asignado[:20] + '...' if len(asignado) > 20 else asignado
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[1*inch, 1*inch, 1*inch, 1*inch, 1.2*inch, 1.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        content.append(table)
    else:
        no_data_para = Paragraph("No se encontraron tickets con los filtros aplicados.", styles['Normal'])
        content.append(no_data_para)
    
    # Construir el PDF
    doc.build(content)
    return response


def generar_excel_reporte(tickets, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia=None):
    """Genera un reporte de tickets en formato Excel"""
    
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Tickets"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Título del reporte
    ws.merge_cells('A1:F2')
    title_cell = ws['A1']
    title_cell.value = "REPORTE DE TICKETS PQRS"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Información de filtros
    row_num = 4
    if fecha_desde or fecha_hasta or (estado and estado != 'todos') or (tipo_solicitud and tipo_solicitud != 'todos'):
        ws.merge_cells(f'A{row_num}:F{row_num}')
        filtros_text = "Filtros aplicados: "
        filtros = []
        if fecha_desde:
            filtros.append(f"Desde: {fecha_desde}")
        if fecha_hasta:
            filtros.append(f"Hasta: {fecha_hasta}")
        if estado and estado != 'todos':
            estado_display = dict(Ticket.ESTADO_CHOICES).get(estado, estado)
            filtros.append(f"Estado: {estado_display}")
        if tipo_solicitud and tipo_solicitud != 'todos':
            tipo_display = dict(Ticket.TIPO_SOLICITUD_CHOICES).get(tipo_solicitud, tipo_solicitud)
            filtros.append(f"Tipo: {tipo_display}")
        
        ws[f'A{row_num}'].value = filtros_text + " | ".join(filtros)
        ws[f'A{row_num}'].font = Font(italic=True)
        row_num += 2
    
    # Encabezados de columnas
    headers = ['Código', 'Tipo', 'Estado', 'Severidad', 'Fecha Creación', 'Asignado a']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for ticket in tickets:
        row_num += 1
        
        # Obtener valores para mostrar
        tipo_display = dict(Ticket.TIPO_SOLICITUD_CHOICES).get(ticket.tipo_solicitud, ticket.tipo_solicitud)
        estado_display = dict(Ticket.ESTADO_CHOICES).get(ticket.estado, ticket.estado)
        severidad_display = dict(Ticket.SEVERIDAD_CHOICES).get(ticket.severidad, ticket.severidad)
        fecha_creacion = ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        asignado = ticket.usuario_asignado.get_full_name() if ticket.usuario_asignado else 'Sin asignar'
        
        # Agregar datos a las celdas
        ws.cell(row=row_num, column=1, value=ticket.codigo)
        ws.cell(row=row_num, column=2, value=tipo_display)
        ws.cell(row=row_num, column=3, value=estado_display)
        ws.cell(row=row_num, column=4, value=severidad_display)
        ws.cell(row=row_num, column=5, value=fecha_creacion)
        ws.cell(row=row_num, column=6, value=asignado)
    
    # Ajustar ancho de columnas
    column_widths = [15, 12, 15, 12, 18, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo
    fecha_actual = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'reporte_tickets_{fecha_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    return response


def generar_txt_reporte(tickets, fecha_desde, fecha_hasta, estado, tipo_solicitud, agencia=None):
    """Genera un reporte de tickets en formato de texto plano"""
    
    # Configurar response para forzar descarga como archivo de texto plano
    response = HttpResponse(content_type='text/plain; charset=utf-8')
    
    # Nombre del archivo con extensión explícita
    fecha_actual = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'reporte_tickets_{fecha_actual}.txt'
    
    # Headers específicos para forzar descarga como archivo TXT
    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    response['Content-Type'] = 'text/plain; charset=utf-8'
    response['Content-Transfer-Encoding'] = 'binary'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    response['X-Content-Type-Options'] = 'nosniff'
    
    # Escribir contenido
    lines = []
    
    # Título
    lines.append("REPORTE DE TICKETS PQRS")
    lines.append("=" * 50)
    lines.append("")
    
    # Información de filtros
    if fecha_desde or fecha_hasta or (estado and estado != 'todos') or (tipo_solicitud and tipo_solicitud != 'todos'):
        lines.append("Filtros aplicados:")
        if fecha_desde:
            lines.append(f"  - Fecha desde: {fecha_desde}")
        if fecha_hasta:
            lines.append(f"  - Fecha hasta: {fecha_hasta}")
        if estado and estado != 'todos':
            estado_display = dict(Ticket.ESTADO_CHOICES).get(estado, estado)
            lines.append(f"  - Estado: {estado_display}")
        if tipo_solicitud and tipo_solicitud != 'todos':
            tipo_display = dict(Ticket.TIPO_SOLICITUD_CHOICES).get(tipo_solicitud, tipo_solicitud)
            lines.append(f"  - Tipo: {tipo_display}")
        lines.append("")
    
    # Información del reporte
    lines.append(f"Fecha de generación: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    lines.append(f"Total de tickets: {tickets.count()}")
    lines.append("")
    
    # Encabezados en formato de texto plano
    lines.append("LISTADO DE TICKETS:")
    lines.append("=" * 80)
    lines.append("")
    lines.append("Código\t\tTipo\t\tEstado\t\tSeveridad\tFecha Creación\t\tAsignado a")
    lines.append("-" * 80)
    
    # Datos
    for ticket in tickets:
        # Obtener valores para mostrar
        tipo_display = dict(Ticket.TIPO_SOLICITUD_CHOICES).get(ticket.tipo_solicitud, ticket.tipo_solicitud)
        estado_display = dict(Ticket.ESTADO_CHOICES).get(ticket.estado, ticket.estado)
        severidad_display = dict(Ticket.SEVERIDAD_CHOICES).get(ticket.severidad, ticket.severidad)
        fecha_creacion = ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        asignado = ticket.usuario_asignado.get_full_name() if ticket.usuario_asignado else 'Sin asignar'
        
        # Formatear datos con espaciado fijo para mejor legibilidad
        row_line = f"{ticket.codigo:<12}\t{tipo_display:<15}\t{estado_display:<12}\t{severidad_display:<10}\t{fecha_creacion:<17}\t{asignado}"
        lines.append(row_line)
    
    # Agregar pie del reporte
    lines.append("")
    lines.append("=" * 80)
    lines.append("Fin del reporte")
    lines.append(f"Generado por: Sistema PQRS - Talento Escucha")
    lines.append(f"Archivo: {filename}")
    lines.append("Formato: Texto plano (.txt)")
    
    # Escribir todo el contenido con codificación explícita
    content = "\n".join(lines)
    response.write(content.encode('utf-8').decode('utf-8'))
    return response


@login_required
@user_has_any_role(['ADMIN', 'REVISOR'])
def perfil_usuario(request):
    """Vista para mostrar y editar el perfil del usuario"""
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu perfil ha sido actualizado exitosamente!')
            return redirect('admin_dashboard:perfil_usuario')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = UserProfileForm(instance=request.user)
    
    # Obtener información adicional del usuario
    user_groups = request.user.groups.all()
    user_roles = [group.name for group in user_groups]
    user_role = user_roles[0] if user_roles else 'Usuario'
    
    # Calcular estadísticas de tickets según el rol del usuario
    if 'ADMIN' in user_roles:
        # Los ADMIN ven estadísticas generales del sistema
        tickets_asignados = Ticket.objects.count()  # Todos los tickets
        tickets_resueltos = Ticket.objects.filter(estado='resuelto').count()  # Todos los resueltos
        tickets_pendientes = Ticket.objects.filter(estado='pendiente').count()  # Todos los pendientes
        
        # También mostrar sus propios tickets asignados si los tiene
        tickets_propios = Ticket.objects.filter(usuario_asignado=request.user).count()
        tickets_propios_resueltos = Ticket.objects.filter(usuario_asignado=request.user, estado='resuelto').count()
        tickets_propios_pendientes = Ticket.objects.filter(usuario_asignado=request.user, estado='pendiente').count()
        
        # Si tiene tickets asignados, mostrar esas estadísticas en su lugar
        if tickets_propios > 0:
            tickets_asignados = tickets_propios
            tickets_resueltos = tickets_propios_resueltos
            tickets_pendientes = tickets_propios_pendientes
            
    else:
        # Los REVISORES ven solo sus tickets asignados
        tickets_asignados = Ticket.objects.filter(usuario_asignado=request.user).count()
        tickets_resueltos = Ticket.objects.filter(usuario_asignado=request.user, estado='resuelto').count()
        tickets_pendientes = Ticket.objects.filter(usuario_asignado=request.user, estado='pendiente').count()
    
    context = {
        'user': request.user,
        'form': form,
        'user_roles': user_roles,
        'user_role': user_role,
        'tickets_asignados': tickets_asignados,
        'tickets_resueltos': tickets_resueltos,
        'tickets_pendientes': tickets_pendientes,
        'last_login': request.user.last_login,
        'date_joined': request.user.date_joined,
    }
    
    return render(request, 'admin_dashboard/perfil_usuario.html', context)


@login_required
@user_has_any_role(['ADMIN', 'REVISOR'])
def cambiar_password(request):
    """Vista para cambiar la contraseña del usuario"""
    
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, '¡Tu contraseña ha sido cambiada exitosamente!')
            return redirect('admin_dashboard:perfil_usuario')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = CustomPasswordChangeForm(request.user)
    
    context = {
        'user': request.user,
        'form': form,
    }
    
    return render(request, 'admin_dashboard/cambiar_password.html', context)
